"""
Imports Letters of Intent from all available source workbooks
(Córdoba, Belém, Illinois) into a single normalized JSON. Replaces
the older import_cordoba_lois.py. Idempotent — re-running on new
workbooks just merges the new rows in.

Sources:
  - data/signed ACE CORDOBA LOI.xlsx                  → Delegates sheet
  - data/signed ACE CORDOBA LOI _ STAKEHOLDERS AND ACADEMIA.xlsx
                                                       → Academia sheet
  - data/ACE Belem LOI signed.xlsx                    → Sheet1
  - data/ACE Illinois LOI data base.xlsx              → bdd sheet
                                                        (cleanest copy)

Output: data/_lois.json — a flat list the observatory can import.
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data"
OUT = DATA / "_lois.json"


def clean(value):
    if value is None:
        return None
    if isinstance(value, str):
        v = re.sub(r"\s+", " ", value).strip()
        return v or None
    return value


def split_pair(text: str | None) -> tuple[str | None, str | None]:
    """Split 'X AND Y' (or variants) into the two halves."""
    if not text:
        return None, None
    parts = re.split(r"\s+AND\s+", text, maxsplit=1, flags=re.IGNORECASE)
    a = clean(parts[0]) if len(parts) >= 1 else None
    b = clean(parts[1]) if len(parts) >= 2 else None
    return a, b


# ── Córdoba (two-party rows, two sheets) ───────────────────────────
def import_cordoba() -> list[dict]:
    out: list[dict] = []
    delegates_path = DATA / "signed ACE CORDOBA LOI.xlsx"
    academia_path = DATA / "signed ACE CORDOBA LOI _ STAKEHOLDERS AND ACADEMIA.xlsx"
    for path, sheet, kind in [
        (delegates_path, "Delegates", "delegates"),
        (academia_path, "Academia", "academia"),
    ]:
        if not path.exists():
            continue
        wb = load_workbook(path, read_only=True, data_only=True)
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        for r in ws.iter_rows(values_only=True):
            if not r or not isinstance(r[0], int):
                continue
            rec = {
                "edition": "ace-22-cordoba-2025",
                "editionLabel": "ACE 22 · Córdoba",
                "kind": kind,
                "partyA": clean(r[1]),
                "countryA": clean(r[2]),
                "partyB": clean(r[3]),
                "countryB": clean(r[4]),
                "purpose": clean(r[5]) if len(r) > 5 else None,
                "delegate": clean(r[6]) if len(r) > 6 else None,
            }
            if rec["partyA"] and rec["partyB"]:
                out.append(rec)
    return out


# ── Belém (single sheet, both parties packed into "X AND Y") ───────
def import_belem() -> list[dict]:
    path = DATA / "ACE Belem LOI signed.xlsx"
    if not path.exists():
        return []
    out: list[dict] = []
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    for r in ws.iter_rows(values_only=True):
        if not r or not isinstance(r[0], int):
            continue
        collab = r[1] if len(r) > 1 else None
        countries = r[2] if len(r) > 2 else None
        purpose = r[3] if len(r) > 3 else None
        rep1 = r[4] if len(r) > 4 else None
        rep2 = r[6] if len(r) > 6 else None
        partyA, partyB = split_pair(collab)
        countryA, countryB = split_pair(countries)
        if partyA and partyB:
            out.append(
                {
                    "edition": "ace-21-belem-2025",
                    "editionLabel": "ACE 21 · Belém",
                    "kind": "delegates",
                    "partyA": partyA,
                    "countryA": countryA.title() if countryA else None,
                    "partyB": partyB,
                    "countryB": countryB.title() if countryB else None,
                    "purpose": clean(purpose),
                    "delegate": clean(rep1) or clean(rep2),
                }
            )
    return out


# ── Illinois (two-party rows, "bdd" sheet has the clean data) ──────
def import_illinois() -> list[dict]:
    path = DATA / "ACE Illinois LOI data base.xlsx"
    if not path.exists():
        return []
    out: list[dict] = []
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet = "bdd" if "bdd" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet]
    header_seen = False
    for r in ws.iter_rows(values_only=True):
        if not r:
            continue
        # Header row contains "Country 1" in the first or second cell.
        if not header_seen:
            if any(c and "Country 1" in str(c) for c in r):
                header_seen = True
            continue
        # Layout: Country1, Stake1, Rep1, Email1, Country2, Stake2, Rep2, Email2, Purpose
        countryA = clean(r[0])
        partyA = clean(r[1])
        rep1 = clean(r[2])
        countryB = clean(r[4]) if len(r) > 4 else None
        partyB = clean(r[5]) if len(r) > 5 else None
        rep2 = clean(r[6]) if len(r) > 6 else None
        purpose = clean(r[8]) if len(r) > 8 else None
        if partyA and partyB:
            out.append(
                {
                    "edition": "ace-20-illinois-2025",
                    "editionLabel": "ACE 20 · Illinois",
                    "kind": "delegates",
                    "partyA": partyA,
                    "countryA": countryA,
                    "partyB": partyB,
                    "countryB": countryB,
                    "purpose": purpose,
                    "delegate": rep1 or rep2,
                }
            )
    return out


def dedupe(records: list[dict]) -> list[dict]:
    seen: set[tuple] = set()
    out: list[dict] = []
    for rec in records:
        key = (
            rec["edition"],
            (rec.get("partyA") or "").lower(),
            (rec.get("partyB") or "").lower(),
            (rec.get("countryA") or "").lower(),
            (rec.get("countryB") or "").lower(),
        )
        if key in seen:
            continue
        seen.add(key)
        out.append(rec)
    return out


def main() -> int:
    cordoba = import_cordoba()
    belem = import_belem()
    illinois = import_illinois()
    merged = dedupe(cordoba + belem + illinois)
    OUT.write_text(
        json.dumps(merged, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    from collections import Counter
    by_edition = Counter(r["edition"] for r in merged)
    print(f"{OUT.name}: wrote {len(merged)} LOIs total")
    for edition, count in by_edition.most_common():
        print(f"  {edition}: {count}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
