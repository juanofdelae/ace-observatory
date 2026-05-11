"""
Imports the 104 Letters of Intent signed at ACE Córdoba 2025 from
the two source Excel files into a single normalized JSON record set
that the observatory can render without re-opening the workbooks.

Sources:
  - data/signed ACE CORDOBA LOI.xlsx                 → Delegates sheet
  - data/signed ACE CORDOBA LOI _ STAKEHOLDERS AND ACADEMIA.xlsx
                                                      → Academia sheet

The two files overlap on the Delegates sheet; we use the smaller
file's Delegates and the second file's Academia, then dedupe by
(partyA + partyB + countryA + countryB).

Output: data/_cordoba-lois.json — a flat list of LOI records the
observatory can import directly.
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data"
F1 = DATA / "signed ACE CORDOBA LOI.xlsx"
F2 = DATA / "signed ACE CORDOBA LOI _ STAKEHOLDERS AND ACADEMIA.xlsx"
OUT = DATA / "_cordoba-lois.json"


def clean(value):
    if value is None:
        return None
    if isinstance(value, str):
        v = re.sub(r"\s+", " ", value).strip()
        return v or None
    return value


def extract(file_path: Path, sheet: str, kind: str) -> list[dict]:
    wb = load_workbook(file_path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        return []
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    out: list[dict] = []
    for r in rows:
        if not r or not isinstance(r[0], int):
            continue
        # Column layout (after the "#"" col):
        #   PARTY A, COUNTRY A, PARTY B, COUNTRY B, DETAIL?, DELEGATE?, EMAIL?
        record = {
            "kind": kind,                              # "delegates" | "academia"
            "edition": "ace-22-cordoba-2025",
            "number": r[0],
            "partyA": clean(r[1]),
            "countryA": clean(r[2]),
            "partyB": clean(r[3]),
            "countryB": clean(r[4]),
            "detail": clean(r[5]) if len(r) > 5 else None,
            "delegate": clean(r[6]) if len(r) > 6 else None,
        }
        if not (record["partyA"] and record["partyB"]):
            continue
        out.append(record)
    return out


def main() -> int:
    delegates = extract(F1, "Delegates", "delegates")
    academia = extract(F2, "Academia", "academia")

    # Dedupe across both sheets by (A, B, countries).
    seen: set[tuple] = set()
    merged: list[dict] = []
    for rec in delegates + academia:
        key = (
            (rec["partyA"] or "").lower(),
            (rec["partyB"] or "").lower(),
            (rec["countryA"] or "").lower(),
            (rec["countryB"] or "").lower(),
        )
        if key in seen:
            continue
        seen.add(key)
        merged.append(rec)

    OUT.write_text(
        json.dumps(merged, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(f"{OUT.name}: wrote {len(merged)} LOIs "
          f"({sum(1 for r in merged if r['kind']=='delegates')} delegates "
          f"+ {sum(1 for r in merged if r['kind']=='academia')} academia)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
