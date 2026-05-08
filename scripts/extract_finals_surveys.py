#!/usr/bin/env python3
"""
Extract structured exit-survey data from the QuestionPro workbooks under
info-reports/FinalsSurvey/ and emit one JSON per edition that the
SurveyDashboard can consume.

The output schema mirrors data/_survey-ace22.json (the existing Córdoba
2025 dashboard data) so the same dashboard component renders both, with
two additional optional sections — `equityRatings` and `connectionsCount`
— plus a `qualitativeQuotes` block for participant comments.

QuestionPro question layout (consistent across the 9 workbooks):
  Q2  [col 23]            overall rating 1–5 (1 = Significantly Above)
  Q3  [cols 24-30]        7 aspect ratings 1–5 (5 = Very Good)
  Q4  [col 31]            recommend (1 = Yes, 2 = Maybe, 3 = No)
  Q5  [cols 32-37]        6 knowledge-topic scores 1–4
  Q6  [cols 38-45]        8 program-impact multi-select flags
  Q9  [col 48]            number of connections made (integer)
  Q13.1 [col 100]         equity statement 1 — 1–5 likert
  Q13.2 [col 101]         equity statement 2 — 1–5 likert
  Q7, Q8, Q12, Q14, Q16   free-text — sampled into qualitativeQuotes
  Country (col 21)        edition-level country distribution

Each output file is data/_survey-ace{N}.json where N is the edition
number. Existing data/_survey-ace22.json is overwritten with the latest
Córdoba data.
"""
from __future__ import annotations
import json
import re
import unicodedata
from collections import Counter
from pathlib import Path
from typing import Any

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "info-reports" / "FinalsSurvey"
OUT_DIR = ROOT / "data"

# Filename prefix → edition id.
EDITION_BY_FILE: dict[str, str] = {
    "Ecuador": "ace-15-ecuador-2022",
    "Louisiana": "ace-14-louisiana-2022",
    "Seattle": "ace-16-seattle-2023",
    "Panama": "ace-17-panama-2024",
    "Michigan": "ace-18-michigan-2024",
    "Armenia": "ace-19-armenia-2024",
    "Illinois": "ace-20-illinois-2025",
    "Belem": "ace-21-belem-2025",
    "Cordoba2025": "ace-22-cordoba-2025",
}

# Q2 overall rating — QuestionPro stores 1 = best.
OVERALL_LABELS = [
    "Significantly Above Expectations",
    "Above Expectations",
    "In Line with Expectations",
    "Below Expectations",
    "Significantly Below Expectations",
]
# Q3 aspect ratings — 5 = Very Good.
ASPECT_LEVELS = ["Very Poor", "Poor", "Neutral", "Good", "Very Good"]
# Q4 recommend.
RECOMMEND_LABELS = ["Yes", "Maybe", "No"]
# Q5 knowledge topic scale — 4 levels.
KNOWLEDGE_LEVELS = ["None", "Low", "Medium", "High"]
# Q13.1/13.2 equity Likert — 1..5.
EQUITY_LEVELS = ["Strongly Disagree", "Disagree", "Neutral", "Agree", "Strongly Agree"]


def num(v: Any) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def cell_at(row: list, idx: int) -> Any:
    """Safe row[idx] — returns None when the column doesn't exist."""
    return row[idx] if 0 <= idx < len(row) else None


def cell_text(v: Any) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    s = unicodedata.normalize("NFC", s)
    s = re.sub(r"\s+", " ", s)
    return s


def distribution(values: list[float | None], levels: list[str]) -> dict:
    """Map numeric scores 1..len(levels) to a labeled distribution +
    arithmetic mean. Levels are in order of ascending score."""
    counts = Counter()
    total = 0
    for lv in levels:
        counts[lv] = 0
    sum_score = 0.0
    for v in values:
        if v is None:
            continue
        idx = int(round(v)) - 1
        if 0 <= idx < len(levels):
            counts[levels[idx]] += 1
            sum_score += v
            total += 1
    options = [
        {
            "label": lv,
            "count": counts[lv],
            "pct": round(counts[lv] / total * 100, 1) if total else 0.0,
        }
        for lv in levels
    ]
    return {
        "options": options,
        "total": total,
        "mean": round(sum_score / total, 2) if total else 0.0,
    }


def aspect_ratings(rows, sub_headers: list[str], cols: range) -> list[dict]:
    """For each Q3 sub-aspect column emit {label, mean, levels}."""
    out = []
    for c in cols:
        label = cell_text(sub_headers[c])
        if not label:
            continue
        scores = [num(r[c]) for r in rows]
        valid = [s for s in scores if s is not None]
        if not valid:
            continue
        bucket = Counter()
        for lv in ASPECT_LEVELS:
            bucket[lv] = 0
        for s in valid:
            idx = int(round(s)) - 1
            if 0 <= idx < len(ASPECT_LEVELS):
                bucket[ASPECT_LEVELS[idx]] += 1
        out.append({
            "label": label,
            "mean": round(sum(valid) / len(valid), 2),
            "levels": {lv: bucket[lv] for lv in ASPECT_LEVELS},
        })
    return out


def knowledge_scope(rows, sub_headers: list[str], cols: range) -> list[dict]:
    """For each Q5 topic column emit {topic, levels}. Knowledge has 4
    levels (None/Low/Medium/High)."""
    out = []
    for c in cols:
        topic = cell_text(sub_headers[c])
        if not topic:
            continue
        scores = [num(r[c]) for r in rows]
        valid = [s for s in scores if s is not None]
        if not valid:
            continue
        bucket = Counter()
        for lv in KNOWLEDGE_LEVELS:
            bucket[lv] = 0
        for s in valid:
            idx = int(round(s)) - 1
            if 0 <= idx < len(KNOWLEDGE_LEVELS):
                bucket[KNOWLEDGE_LEVELS[idx]] += 1
        out.append({
            "topic": topic,
            "exit": {lv: bucket[lv] for lv in KNOWLEDGE_LEVELS},
            # No pre-survey for these editions — keep the field for schema
            # parity but fill with zeros so the dashboard knows there's
            # no growth comparison.
            "pre": {lv: 0 for lv in KNOWLEDGE_LEVELS},
        })
    return out


def program_impact(rows, sub_headers: list[str], cols: range) -> list[dict]:
    """Q6 multi-select — count how many respondents picked each flag."""
    out = []
    total_resp = sum(1 for r in rows if any(num(r[c]) is not None for c in cols))
    for c in cols:
        label = cell_text(sub_headers[c])
        if not label:
            continue
        count = sum(1 for r in rows if num(r[c]) is not None)
        out.append({
            "label": label,
            "count": count,
            "pct": round(count / total_resp * 100, 1) if total_resp else 0.0,
        })
    return out


def country_distribution(rows, country_col: int) -> list[dict]:
    cnt = Counter()
    for r in rows:
        c = cell_text(r[country_col])
        if c:
            cnt[c] += 1
    total = sum(cnt.values())
    return [
        {
            "country": k,
            "count": v,
            "pct": round(v / total * 100, 1) if total else 0.0,
        }
        for k, v in sorted(cnt.items(), key=lambda kv: -kv[1])
    ]


def connections_summary(rows, col: int) -> dict:
    vals = [num(r[col]) for r in rows]
    valid = [v for v in vals if v is not None]
    if not valid:
        return {"mean": 0.0, "median": 0, "total": 0, "buckets": []}
    valid_int = [int(round(v)) for v in valid]
    valid_sorted = sorted(valid_int)
    median = valid_sorted[len(valid_sorted) // 2]
    # Buckets for histogram-style display.
    bucket_specs = [(0, 0), (1, 2), (3, 5), (6, 10), (11, 9999)]
    bucket_labels = ["0", "1–2", "3–5", "6–10", "11+"]
    buckets = []
    for (lo, hi), lab in zip(bucket_specs, bucket_labels):
        c = sum(1 for v in valid_int if lo <= v <= hi)
        buckets.append({"range": lab, "count": c})
    return {
        "mean": round(sum(valid_int) / len(valid_int), 2),
        "median": median,
        "total": sum(valid_int),
        "buckets": buckets,
    }


def equity_ratings(rows, q131_col: int, q132_col: int) -> dict:
    return {
        "genderYouth": distribution([num(r[q131_col]) for r in rows], EQUITY_LEVELS),
        "equityImportance": distribution([num(r[q132_col]) for r in rows], EQUITY_LEVELS),
    }


def qualitative_quotes(rows, headers: list[str], max_per_q: int = 5) -> list[dict]:
    """Pick up to `max_per_q` non-trivial free-text answers per
    qualitative question, with attribution."""
    qcols = []
    for i, h in enumerate(headers):
        if not h:
            continue
        h = str(h)
        if h.startswith(("7.", "8.", "12.", "14.", "16.")):
            qcols.append((i, h.split(" ", 1)[0].rstrip(".")))
    out = []
    for col, qid in qcols:
        picks = []
        for r in rows:
            text = cell_text(r[col])
            if len(text) < 50 or len(text) > 600:
                continue
            picks.append({
                "question": qid,
                "name": cell_text(r[18]) or "Anonymous",
                "country": cell_text(r[21]) or cell_text(r[16]),
                "text": text,
            })
            if len(picks) >= max_per_q:
                break
        out.extend(picks)
    return out


def detect_edition_for(filename: str) -> str | None:
    base = filename.split("-")[0].strip()
    return EDITION_BY_FILE.get(base)


def process_workbook(path: Path) -> dict | None:
    edition_id = detect_edition_for(path.name)
    if not edition_id:
        print(f"  ! skipped (unknown edition): {path.name}")
        return None
    wb = openpyxl.load_workbook(path, data_only=True)
    if "Raw Data" not in wb.sheetnames:
        print(f"  ! no Raw Data sheet: {path.name}")
        return None
    ws = wb["Raw Data"]
    headers = [c.value for c in ws[1]]
    sub_headers = [c.value for c in ws[2]]
    raw_rows = [list(row) for row in ws.iter_rows(min_row=3, values_only=True)]
    # Pad every row up to `WIDE` columns so r[N] never IndexErrors when a
    # workbook (e.g. ACE Belem) has fewer Q10 host-site sub-columns than
    # the QuestionPro template max. We sample a generous width — beyond
    # the highest column we read.
    WIDE = 110
    rows = [
        list(r) + [None] * max(0, WIDE - len(r))
        for r in raw_rows
        if any(v not in (None, "") for v in r)
    ]
    sub_headers = list(sub_headers) + [None] * max(0, WIDE - len(sub_headers))
    headers = list(headers) + [None] * max(0, WIDE - len(headers))
    total = len(rows)

    survey = {
        "editionId": edition_id,
        "totalResponses": total,
        "overallRating": distribution([num(r[23]) for r in rows], OVERALL_LABELS),
        "aspectRatings": aspect_ratings(rows, sub_headers, range(24, 31)),
        "recommend": distribution([num(r[31]) for r in rows], RECOMMEND_LABELS),
        "knowledgeScope": knowledge_scope(rows, sub_headers, range(32, 38)),
        "programImpact": program_impact(rows, sub_headers, range(38, 46)),
        "connectionsCount": connections_summary(rows, 48),
        "equityRatings": equity_ratings(rows, 100, 101),
        "countryDistribution": country_distribution(rows, 21),
        "qualitativeQuotes": qualitative_quotes(rows, headers),
        "knowledgeGrowth": [],  # no pre-survey for these editions
        "sessionsAttended": [],  # not asked in QuestionPro template
    }
    return survey


def main() -> None:
    files = sorted(SRC.glob("*.xlsx"))
    if not files:
        print(f"No xlsx files in {SRC}")
        return
    written = 0
    for f in files:
        print(f"--- {f.name}")
        s = process_workbook(f)
        if not s:
            continue
        n = s["editionId"].split("-")[1]
        out = OUT_DIR / f"_survey-ace{n}.json"
        out.write_text(json.dumps(s, indent=2, ensure_ascii=False) + "\n")
        print(f"  → {out.name}: {s['totalResponses']} responses, "
              f"overall mean {s['overallRating']['mean']}, "
              f"{len(s['qualitativeQuotes'])} quotes")
        written += 1
    print(f"\nWrote {written} survey JSONs.")


if __name__ == "__main__":
    main()
