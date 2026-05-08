#!/usr/bin/env python3
"""
Pulls structured exit-survey data out of the ACE Córdoba (ACE 22) Excel
workbooks shipped under info-reports/ and writes
data/_survey-ace22.json — a small, UI-ready bundle that the dashboard
consumes through `data/surveys.ts`.

Sources:
  - ACE Cordoba Exit Survey_nov19.xlsx   (26 responses, latest)
  - Book1.xlsx                           (pre-survey vs exit-survey
                                          knowledge-level comparison
                                          per topic)
"""
from __future__ import annotations
import json
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SURVEY_FILE  = ROOT / "info-reports" / "ACE Cordoba Exit Survey_nov19.xlsx"
COMPARE_FILE = ROOT / "info-reports" / "Book1.xlsx"
OUT_JSON     = ROOT / "data" / "_survey-ace22.json"


def parse_overall_rating(ws) -> dict:
    """Sheet Q3-Q7 starts with the overall-rating block."""
    rating_options = {
        "Significantly Above Expectations", "Above Expectations",
        "In Line with Expectations", "Below Expectations",
        "Significantly Below Expectations",
    }
    out = []
    total = 0
    for row in ws.iter_rows(values_only=True):
        if not row or not row[0]:
            continue
        label = str(row[0]).strip()
        if label in rating_options and row[1] is not None:
            count = int(row[1])
            pct = float(row[2]) if row[2] is not None else 0.0
            out.append({"label": label, "count": count, "pct": round(pct * 100, 1)})
            total += count
        elif label.startswith("Total") and out and total == 0:
            total = int(row[1] or 0)
        elif label.startswith("Mean") and out:
            return {"options": out, "total": total, "mean": float(row[1] or 0)}
    return {"options": out, "total": total, "mean": 0.0}


def parse_aspect_ratings(ws) -> list[dict]:
    """The 7-aspect block follows. Each aspect has its own Very Poor … Very Good
    distribution + a Mean line. We harvest aspect → mean (5-point scale)."""
    aspects = []
    capture_after = False
    current_aspect = None
    levels = {}
    rows = list(ws.iter_rows(values_only=True))
    section_aspects = {"Overall agenda", "Presentations", "Group of participants",
                       "Logistics", "Hotels", "Topics/theme", "Support team"}
    rating_levels = ["Very Poor", "Poor", "Neutral", "Good", "Very Good"]

    for i, row in enumerate(rows):
        if not row or not row[0]:
            continue
        label = str(row[0]).strip()
        if label in section_aspects:
            current_aspect = label
            levels = {}
            continue
        if current_aspect:
            if label in rating_levels and row[1] is not None:
                levels[label] = int(row[1])
            elif label == "Mean" and row[1] is not None:
                aspects.append({
                    "label": current_aspect,
                    "mean": round(float(row[1]), 2),
                    "levels": levels,
                })
                current_aspect = None
                levels = {}
    return aspects


def parse_recommend(ws) -> dict:
    """Q4: 'Would you recommend the ACE Program to other leaders?'"""
    rows = list(ws.iter_rows(values_only=True))
    yes = no = 0
    for i, row in enumerate(rows):
        if not row or not row[0]:
            continue
        label = str(row[0]).strip()
        if label == "Yes" and row[1] is not None:
            yes = int(row[1])
        elif label == "No" and row[1] is not None:
            no = int(row[1])
            break
    total = yes + no
    return {
        "yes": yes,
        "no": no,
        "yesPct": round(yes / total * 100, 1) if total else 0,
    }


def parse_knowledge_scope(ws) -> list[dict]:
    """Q5: knowledge level (None/Low/Medium/High) per topic block (5 topics)."""
    topics = {"Agribusiness & AgriTech", "Automotive & Auto Parts Industry",
              "Connectivity & Digital Transformation",
              "Startup & Innovation Ecosystem",
              "Life Sciences: HealthTech & BioTech"}
    levels = ["None", "Low", "Medium", "High"]
    out = []
    rows = list(ws.iter_rows(values_only=True))
    current = None
    by_level: dict[str, int] = {}
    for row in rows:
        if not row or not row[0]:
            continue
        label = str(row[0]).strip()
        if label in topics:
            current = label
            by_level = {}
            continue
        if current:
            for lvl in levels:
                if label == lvl and row[1] is not None:
                    by_level[lvl] = int(row[1])
                    break
            if label == "Mean" and row[1] is not None:
                out.append({
                    "topic": current,
                    "mean": round(float(row[1]), 2),
                    "distribution": by_level,
                })
                current = None
                by_level = {}
    return out


def parse_program_impact(ws) -> list[dict]:
    """Q6: 'How did ACE Córdoba impact you?' multi-select counts."""
    impacts: list[dict] = []
    in_block = False
    rows = list(ws.iter_rows(values_only=True))
    for row in rows:
        if not row:
            continue
        label = str(row[0] or "").strip()
        if label.startswith("6. Please indicate how the ACE program"):
            in_block = True
            continue
        if not in_block:
            continue
        if label == "Total":
            break
        if not label or label in ("Other Option [Other]", "Response ID"):
            continue
        if row[1] is None:
            continue
        impacts.append({"label": label, "count": int(row[1])})
    return impacts


def parse_session_count(ws) -> list[dict]:
    """Q9 (sheet 'Q10 - Q12'): how many sessions did the participant attend?"""
    out = []
    rows = list(ws.iter_rows(values_only=True))
    in_block = False
    for row in rows:
        if not row: continue
        label = str(row[0] or "").strip()
        if label.startswith("9. Please indicate the number"):
            in_block = True
            continue
        if not in_block: continue
        if label == "Total": break
        if not label or row[1] is None: continue
        if any(c.isdigit() for c in label) or label in ("0", "61+"):
            out.append({"range": label, "count": int(row[1])})
    return out


def parse_pre_vs_exit(wb: openpyxl.Workbook) -> list[dict]:
    """
    Book1.xlsx layout:
        Row N    : <Topic name>  Pre-Survey Pre-Survey Exit Survey Exit Survey   (only the FIRST topic has these column headers)
        Row N+1  : None  4 0.14 1 0.04
        Row N+2  : Low   12 0.43 0 0
        ... ↓ Total → next topic begins
        Row M    : <NextTopic>      (blank columns afterwards)
        Row M+1  : None  ...        (level rows pick up from here)
    We keep a "current topic" cursor and reset it when we see either a
    bare topic-name row (no level words present) OR a Total marker.
    """
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    out: list[dict] = []
    current = None
    pre: dict[str, int] = {}
    exit_: dict[str, int] = {}
    LEVELS = {"None", "Low", "Medium", "High"}

    def flush():
        nonlocal current, pre, exit_
        if current and pre and exit_:
            out.append({"topic": current, "pre": dict(pre), "exit": dict(exit_)})
        current, pre, exit_ = None, {}, {}

    for row in rows:
        if not row:
            continue
        label = str(row[0] or "").strip()
        if not label:
            continue
        bare = label.replace("\xa0", "").strip()

        if bare in LEVELS and current:
            try:
                pre_count  = int(row[1]) if row[1] is not None else 0
                exit_count = int(row[3]) if row[3] is not None else 0
            except (TypeError, ValueError):
                pre_count = exit_count = 0
            pre[bare] = pre_count
            exit_[bare] = exit_count
            continue

        if label.startswith("Total"):
            flush()
            continue

        # Anything else with text is a new topic header — start a fresh block.
        if bare not in LEVELS:
            flush()
            current = label
    flush()
    return out


def main() -> int:
    if not SURVEY_FILE.exists() or not COMPARE_FILE.exists():
        print("missing source xlsx", file=sys.stderr)
        return 1

    survey_wb = openpyxl.load_workbook(SURVEY_FILE, data_only=True)
    q3_q7 = survey_wb["Q3 - Q7"]
    q10_q12 = survey_wb["Q10 - Q12"]
    overall = parse_overall_rating(q3_q7)
    aspects = parse_aspect_ratings(q3_q7)
    recommend = parse_recommend(q3_q7)
    knowledge = parse_knowledge_scope(q3_q7)
    impact = parse_program_impact(q3_q7)
    sessions = parse_session_count(q10_q12)

    compare_wb = openpyxl.load_workbook(COMPARE_FILE, data_only=True)
    pre_vs_exit = parse_pre_vs_exit(compare_wb)

    # Country breakdown from Q2 contact info.
    countries: dict[str, int] = {}
    q2 = survey_wb["Q2"]
    started = False
    for row in q2.iter_rows(values_only=True):
        if not row: continue
        if row[0] and str(row[0]).strip().isdigit():
            country = (row[5] or "").strip() if len(row) > 5 else ""
            if country:
                countries[country] = countries.get(country, 0) + 1
    country_list = sorted(
        [{"country": k, "count": v} for k, v in countries.items()],
        key=lambda x: -x["count"],
    )

    out = {
        "editionId": "ace-22-cordoba-2025",
        "totalResponses": overall.get("total", 0),
        "overallRating": overall,
        "aspectRatings": aspects,
        "recommend": recommend,
        "knowledgeScope": knowledge,
        "programImpact": impact,
        "sessionsAttended": sessions,
        "knowledgeGrowth": pre_vs_exit,
        "countryDistribution": country_list,
    }
    OUT_JSON.write_text(json.dumps(out, ensure_ascii=False, indent=2))
    print(f"Wrote {OUT_JSON.relative_to(ROOT)}")
    print(f"  total responses: {out['totalResponses']}")
    print(f"  aspects: {len(aspects)}")
    print(f"  knowledge topics: {len(knowledge)}")
    print(f"  pre vs exit topics: {len(pre_vs_exit)}")
    print(f"  countries: {len(country_list)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
